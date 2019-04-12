import openpyxl
import os
from tkinter import messagebox

filename = 'account_details.xlsx'

if os.path.isfile(filename):
    wb = openpyxl.Workbook(filename)
    if 'Data' in wb.get_sheet_names():
        pass
    else:
        wb.create_sheet(index=0, title='Data')
else:
   wb = openpyxl.Workbook()
   wb.create_sheet(index=0, title='Data')
   wb.save(filename)

wb = openpyxl.load_workbook(filename)
sheet = wb.get_sheet_by_name('Data')
print("workbook")

def add_entry(payee_name, payee_account_num, payee_ifsc, payee_bank, payee_address, payee_amount, payee_amount_words, payee_date):
    ws = wb.active
    if payee_name == "" or payee_ifsc == "" or payee_amount == "":
        messagebox.showerror("Error", "Payee details Insufficient")
    column1 = ws['A']
    column2 = ws['B']
    column3 = ws['C']
    column4 = ws['D']
    column5 = ws['E']
    column6 = ws['F']
    column7 = ws['G']
    column8 = ws['H']

    col_len1 = str(len(column1)+1)
    col_len2 = str(len(column2)+1)
    col_len3 = str(len(column3)+1)
    col_len4 = str(len(column4)+1)
    col_len5 = str(len(column5)+1)
    col_len6 = str(len(column6)+1)
    col_len7 = str(len(column7)+1)
    col_len8 = str(len(column8)+1)


    sheet['A' + col_len1] = payee_name
    sheet['B' + col_len2] = payee_account_num
    sheet['C' + col_len3] = payee_ifsc
    sheet['D' + col_len4] = payee_bank
    sheet['E' + col_len5] = payee_address
    sheet['F' + col_len6] = payee_amount
    sheet['G' + col_len7] = payee_amount_words
    sheet['H' + col_len8] = payee_date
    wb.save(filename)

    if ((sheet['A1'].value == 'payee') and
        (sheet['B1'].value == 'AccountNumber') and
        (sheet['C1'].value == 'ifsc') and
        (sheet['D1'].value == 'bank') and
        (sheet['E1'].value == 'address') and
        (sheet['F1'].value == 'Amount') and
        (sheet['G1'].value == 'Amountinwords') and
        (sheet['H1'].value == 'date')):
            pass
    else:
        sheet['A1'].value ='payee'
        sheet['B1'].value = 'AccountNumber'
        sheet['C1'].value = 'ifsc'
        sheet['D1'].value = 'bank'
        sheet['E1'].value = 'address'
        sheet['F1'].value = 'Amount'
        sheet['G1'].value = 'Amountinwords'
        sheet['H1'].value = 'date'

import json
from mailmerge import MailMerge
from datetime import date

template = "template.docx"

def add_to_json(payee_name, payee_account_num, payee_ifsc, payee_bank, payee_address, payee_amount, payee_amount_words, payee_date):
    with open('data.json') as inFile:
        try:
            data = json.load(inFile)
        except ValueError:
             data = {}
        data[payee_name] = {
        'payee_name': payee_name,
        'payee_acc': payee_account_num,
        'payee_ifsc': payee_ifsc,
        'payee_bank': payee_bank,
        'payee_bank_add':payee_address,
        'amount':payee_amount,
        'amount_in_words':payee_amount_words,
        'date':payee_date
        }

    with open('data.json', 'w') as outfile:
        json.dump(data, outfile)

def search_entry_and_generate_doc(name):
    print(name)
    document = MailMerge(template)
    print(document.get_merge_fields())
    with open('data.json') as infile:
        try:
            data = json.load(infile)
            print(data[name])
            output = data[name]
            document.merge(**output)
            document.write("output.docx")
            os.startfile("output.docx", "print")
        except Exception as e:
            print(e)
            messagebox.showerror("Error","File doesnot have any record")

